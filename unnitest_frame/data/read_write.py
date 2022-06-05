
import openpyxl

class ReadWrite:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet
        self.wb=openpyxl.load_workbook(self.file)
        self.table=self.wb[self.sheet]
        self.table=self.wb.active
        self.nrow=self.table.max_row        #获取最大行
        self.ncol=self.table.max_column     #获取最大列

    def read(self):
        list2=[]
        for row in range(2,self.nrow+1):
            list1=[]
            for column in range(1,self.ncol+1):
                content=self.table.cell(row,column).value
                list1.append(content)
            list2.append(list1)
        self.wb.close()
        return list2

    def write(self,*arg):
        for col in range(1,len(arg)+1):
            self.table.cell(self.nrow+1,col).value=arg[col-1] #将arg元组的元素按索引分别填入excel表里
        self.wb.save(self.file)
        self.wb.close()


if __name__ =="__main__":
    file=r"D:\Python\loguru\data\testdata.xlsx"
    sheet="login"
    doc1=ReadWrite(file,sheet)
    print(doc1.read())